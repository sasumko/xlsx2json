
import sys
import os
import json

from jkLib import jkCommonLib

from openpyxl import Workbook
from openpyxl import load_workbook

gListTextSheets = []

gOutputPath = ""
gOutputHeaderPath = ""

def GetLastKeyColumnIndex (_dict):
    _ret = 0
    for _key, _list in _dict.items():
        for _columnId in _list:
            if _columnId > _ret:
                _ret = _columnId
    return _ret

def GetKeyByColumnIndex (_dict, _columnIndex):
    for _key, _list in _dict.items():
        for _columnId in _list:
            if _list.count(_columnIndex) > 0:
                return _key
    return ""

def GetKeyFieldStartRowIndex (_sheet):
    ## fine next row that its first cell filled
    _rowIndex = 3
    for _row in _sheet.iter_rows(min_row=3, max_row=9999):
        if _row[0].value == None:
            _rowIndex = _rowIndex + 1
            continue
        break
    return _rowIndex


def ParseKeyField (_sheet, _keyRowIndex):
    ## read field
    _dictRet = {}
    _columnIndex = 0    #A = 0

    for _field in _sheet[_keyRowIndex]:
        _strField = GetMergedCellValue(_sheet, _keyRowIndex, _columnIndex)
        if _strField == None:
            if _field.value == None:
                break
            _strField = str.strip(_field.value)
        
        if jkCommonLib.IsEmpty(_strField) == True or _strField.startswith('_'):
            _columnIndex = _columnIndex + 1
            continue
        
        if _dictRet.get(_strField) == None:
            _dictRet[_strField] = []
        
        _listColumn = _dictRet[_strField]

        if _listColumn == None:
            #Add it!
            _dictRet[_strField] = [ _columnIndex ]
        else:
            _dictRet[_strField].append(_columnIndex)
        _columnIndex = _columnIndex + 1    

    _listField = list(_dictRet.keys())
    _strFields = "......> : "

    print ('.................................................')
    print ('......>key field parse done')
    
    for _field in _listField:
        _strFields = _strFields + _field + " "

    print (_strFields)

    return _dictRet 


def GetMergedCellValue (_sheet, _row, _col):
    for e in _sheet.merged_cells:
        if _row != e.min_row:
            continue
        if _row != e.max_row:
            continue
        
        if e.min_col > _col:
            continue

        if e.max_col < _col:
            continue

        return _sheet.cell(e.min_row, e.min_col).value    
    return None

def WriteTextHeader (_textSheetNames):
    _lines = 'public class H_Str\n{\n'
    _lines += '\tpublic string[] TextFileNames = {'

    for index, _name in enumerate( _textSheetNames ):
        _lines += '\n\t\t\"%s\"' % _name
        if index < len(_textSheetNames) - 1:
            _lines += ','
    _lines += '\n\t};'
    
    _lines += '\n\n\tpublic enum eTab\n\t{'
    for index, _name in enumerate( _textSheetNames ):
        _lines += '\n\t\t%s' % _name
        if index == 0:
            _lines += ' = 0'
        if index < len(_textSheetNames) - 1:
            _lines += ','
    _lines += '\n\t};'

    _lines += '\n}'
    print (_lines)
    return _lines

def ParseSheetSubProcHeader(_sheet, _name):
    ####
    ##  Parse sheet as header

    _dataRowIndex = GetKeyFieldStartRowIndex(_sheet)
    
    _lines = 'public class %s\n{\n' % jkCommonLib.GetGNUString( _name )

    for _row in _sheet.iter_rows(min_row=_dataRowIndex + 1, max_col=2):
        if _row[0].value == None:
            continue

        _str_key = jkCommonLib.GetGNUString(_row[0].value) #str.strip(_row[0].value).upper()
        _str_value = _row[1].value

        _str_line_type = ''
        if jkCommonLib.IsInt(_str_value) == True:
            _str_line_type = 'int'
            
        elif jkCommonLib.IsFloat(_str_value) == True:
            _str_line_type = 'float'
            _str_value = '%sf' % _str_value
            
        else:
            _str_line_type = 'string'
            _str_value = '\"%s\"' % _str_value
            
        _lines = _lines + '\tpublic const %s\t%s\t=\t%s;\n' % (_str_line_type, _str_key, _str_value)


    _lines = _lines +  '\n}'
    print ( _lines)
    
    return _lines

gStrLanguages = { 'ko', 'en', 'ja', 'zh-tw', 'zh-ch' }

def ParseSheetSubProcText (_sheet):
    ####
    ##  Parse sheet as Text
    _dataRowIndex = GetKeyFieldStartRowIndex(_sheet)
    _dictFieldByColumn = ParseKeyField(_sheet, _dataRowIndex)
    _lastColumnIndex = GetLastKeyColumnIndex(_dictFieldByColumn)

    _dictRet = {}   #return value (dict)

    
    print ('.................................................')
    ##  read data by field order
    for _row in _sheet.iter_rows(min_row=_dataRowIndex + 1, max_col=_lastColumnIndex + 1):
        if _row[0].value == None:
            #if first column is none == comment row to be skipped
            continue

        _columnIndex = 0
        

        _idToStore = ""
        _dictStr = {}
        
        #read row
        for _value in _row:
            _key = GetKeyByColumnIndex(_dictFieldByColumn, _columnIndex)
            _valueToStore = _value.value

            if jkCommonLib.IsEmpty(_key) == True or _valueToStore == None:
                _columnIndex = _columnIndex + 1
                continue
            
            if _key == 'ID':
                _idToStore = _valueToStore

            if len(_idToStore) > 0 and _key in gStrLanguages:
                # _key => LANGUAGE CODE
                if _key not in _dictRet:
                    _dictRet[_key] = {}

                if _idToStore in _dictRet[_key]:
                    print ("Error!! KEY %s DUPLICATED! " % (_idToStore ))
                _dictRet[_key][_idToStore] = _valueToStore
                # _dictRow = _dictRet[_key]
                
                # _add = {}
                # _add[_idToStore] = _valueToStore
                
                # _dictRow.append(_add)
                # _listRow.append(_add)
            
            _columnIndex = _columnIndex + 1

    _strToDump = json.dumps(_dictRet)
    print (_strToDump)
    return _dictRet

def ParseSheetSubProcJsonData(_sheet):
    ####
    ##  Parse sheet as Json data table
    #for e in _sheet.merged_cells:
        #print (e)
        #print (_sheet.cell(e.min_row, e.min_col).value)
        # _base_value = _sheet.cell_value(
        # print (_base_value)

    _dataRowIndex = GetKeyFieldStartRowIndex(_sheet)
    _dictFieldByColumn = ParseKeyField(_sheet, _dataRowIndex)
    _lastColumnIndex = GetLastKeyColumnIndex(_dictFieldByColumn)

    _listRet = []   #return value



    print ('.................................................')
    ##  read data by field order
    for _row in _sheet.iter_rows(min_row=_dataRowIndex + 1, max_col=_lastColumnIndex + 1):
        if _row[0].value == None:
            #if first column is none == comment row to be skipped
            continue
        
        _columnIndex = 0
        _dictRow = {}

        print ('......> --- [%d] --- ' % len(_listRet) )
        for _value in _row:
            _key = GetKeyByColumnIndex(_dictFieldByColumn, _columnIndex)
            _valueToStore = _value.value
            if jkCommonLib.IsEmpty(_key) == True or _valueToStore == None:
                _columnIndex = _columnIndex + 1
                continue
            
            if _dictRow.get(_key) == None:
                #Add value as single data
                _dictRow[_key] = _valueToStore
                print('......>> [%s] = [%s]' % (_key, _valueToStore))
            else:
                if isinstance( _dictRow[_key], (list,)) == True:
                    _dictRow[_key].append( _valueToStore )
                    print('......>> [%s] = %s (array updated)' % (_key, _dictRow[_key]))
                else:
                    _dictRow[_key] = [_dictRow[_key], _valueToStore]
                    print('......>> [%s] = %s (array updated)' % (_key, _dictRow[_key]))
                # print ('key:%s\t\tvalue:%s' % (_key, _valueToStore))

            _columnIndex = _columnIndex + 1
        
        if len( _dictRow.keys() ) > 0:
            _listRet.append(_dictRow)
        

    _strToDump = json.dumps(_listRet)
    return _listRet


def ParseSheet (_sheet):
    if jkCommonLib.IsEmpty(_sheet['A1'].value) == True or jkCommonLib.IsEmpty(_sheet['A2'].value):
        return None
        
    _nameCellValue = str.strip( _sheet['A1'].value ).lower()
    _typeCellValue = str.strip( _sheet['A2'].value ).lower()
    
    # is valid to parse?
    if _nameCellValue != "name":
        return None

    if _typeCellValue != "type":
        return None

    ##  read def
    _sheetName = _sheet.title
    _outputFileName = str.strip(_sheet['B1'].value)
    
    _type = str.strip(_sheet['B2'].value).lower()
    print ('.................................................')
    print ('...start to parse sheet \"%s\"...' % _sheetName)
    print ('......>output name = %s' % _outputFileName)
    print ('......>Type = %s' % _type)

    _listRet = []
    if _type == 'table':
        _listRet = ParseSheetSubProcJsonData(_sheet)
        if _listRet == None or len(_listRet) == 0:
            print ('.................................................')
            print ('......> Nothing to parse!')
            print ('.................................................')
            print ('\n\n')
            
            return None
    
        _outputFileNameWithPath = os.path.join( gOutputPath, _outputFileName + '.json')
        _fpJson = open(_outputFileNameWithPath, 'w')
        json.dump(_listRet, _fpJson)
        _fpJson.close()
        print ('.................................................')
        print ('......> Write file [%s] done! ' % _outputFileNameWithPath)
        print ('.................................................')
        print ('\n\n')
    elif _type == 'header':
        #ParseSheet as header
        _strToWrite = ParseSheetSubProcHeader(_sheet, _outputFileName)

        WriteHeaderFile(_outputFileName, _strToWrite)

    
    elif _type == 'text':
        #parse sheet as text
        _dictRet = ParseSheetSubProcText(_sheet)

        

        
        for _lang, _list in _dictRet.items():
            _pathLang = os.path.join( gOutputPath, _lang)
            if not os.path.exists(_pathLang):
                os.makedirs(_pathLang)
            
            #write files to each language's directory
            _outputFileNameWithPath = os.path.join( _pathLang, _outputFileName + ".json")
            _fpJson = open(_outputFileNameWithPath, 'w', encoding='utf8')
            json.dump(_list, _fpJson, indent=3, ensure_ascii=False )
            _fpJson.close()
            if _outputFileName not in gListTextSheets:
                gListTextSheets.append(_outputFileName)
        
        
        
                
            
    return _listRet
            


def WriteHeaderFile (_fileName, _buf):
    _outputFileNameWithPath = os.path.join( gOutputHeaderPath, _fileName + '.cs')
    _fpHeader = open(_outputFileNameWithPath, 'w')
    _fpHeader.write(_buf)
    _fpHeader.close()

def xlsx2json (_sourceFile):
    _workbook = load_workbook(_sourceFile, data_only=True)

    print (_workbook.sheetnames)

    for _sheet in _workbook:
        ParseSheet(_sheet)

    #in case that Text parser executed, need to write down string header
    if len( gListTextSheets ) > 0:
        _lineToWrite = WriteTextHeader(gListTextSheets)
        WriteHeaderFile("H_Str", _lineToWrite)
    return 1

_debugFlag = 0

def main (argv = sys.argv):
    global gOutputPath
    global gOutputHeaderPath
    if _debugFlag == 0:
        if len (argv) < 2:
            print ('usage : xlsx2json [xlsx file] [output]')
            return -1

        _xlsFileName = jkCommonLib.GetFullPath(argv[1])
        
        if len(argv) > 2:
            gOutputPath = jkCommonLib.GetFullPath(argv[2])
            gOutputHeaderPath = gOutputPath
            
            if len(argv) > 3:
                gOutputHeaderPath = jkCommonLib.GetFullPath(argv[3])
            #print (_outputHeaderPath)
        else:
            gOutputPath = jkCommonLib.GetLocatedPath(_xlsFileName)
            gOutputHeaderPath = gOutputPath

    else:
    
        _xlsFileName = jkCommonLib.GetFullPath("../../sample.xlsx")
        gOutputPath = jkCommonLib.GetLocatedPath(_xlsFileName)
        gOutputHeaderPath = gOutputPath

    if os.path.isfile(_xlsFileName) == False:
        print ("Error: Cannout find file %s!" % _xlsFileName)
    
    print ('start reading file %s on @%s' 
            % (jkCommonLib.GetFileNameFromPath(_xlsFileName), 
                jkCommonLib.GetLocatedPath(_xlsFileName) )
        )
    print ('export json file on @%s' % gOutputPath )
    
    
    xlsx2json(_xlsFileName)
    

    return 1

if __name__ == "__main__":
    ret = main()
    # sys.exit(ret)

